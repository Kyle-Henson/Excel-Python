import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import argparse

# This script will append all cells in column 3 or greater to a new row while keeping the first two columns

parser = argparse.ArgumentParser()
parser.add_argument("--input", "-i", type=str, required=True)
parser.add_argument("--output", "-o", type=str, required=True)
args = parser.parse_args()

read_file = args.input
write_file = args.output

#read_file = 'G:\\Document.xlsx'
#write_file = 'G:\\Document-Updated.xlsx'

wb = openpyxl.load_workbook(filename=read_file,
     read_only=True)
ws = wb.get_sheet_by_name('Sheet1')

wb2 = openpyxl.Workbook()
ws2 = wb2.create_sheet()

def append_row(row_num,colA,colB,colC):
    ws2.cell(row=row_num, column=1).value = str(colA)
    ws2.cell(row=row_num, column=2).value = str(colB)
    ws2.cell(row=row_num, column=3).value = str(colC)
    print(row_num, colA, colB, colC)

row_num = 1
for row in ws.iter_rows(min_row=row_num):
    # reset the column on a new row
    col_counter = 1
    for cell in row:
        if cell.value:
            if col_counter == 1:
                colA = cell.value
            elif col_counter == 2:
                colB = cell.value
            # Every column greater than or equal to 3 will be added to a new row including the same first 2 columns
            elif col_counter >= 3:
                colC = cell.value
                append_row(row_num, colA, colB, colC)
                # iterate the writing row number
                row_num += 1
            # iterate the current row column
            col_counter += 1

wb2.save(write_file)
wb.close()
wb2.close()




