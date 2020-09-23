import openpyxl
from openpyxl.styles import PatternFill

# Colors a heatmap based on values 1-4

wb = openpyxl.load_workbook(
    'C:\\Document.xlsx',
     data_only=True)

# RGB color picker https://www.rapidtables.com/web/color/RGB_Color.html
red_1 = PatternFill(start_color='C00000',
                    end_color='C00000',
                    fill_type='solid')

red_125 = PatternFill(start_color='D8230A',
                      end_color='D8230A',
                      fill_type='solid')

red_15 = PatternFill(start_color='FF7C80',
                     end_color='FF7C80',
                     fill_type='solid')

red_175 = PatternFill(start_color='FF9B9B',
                      end_color='FF9B9B',
                      fill_type='solid')

orange_2 = PatternFill(start_color='ED7D31',
                       end_color='ED7D31',
                       fill_type='solid')

orange_225 = PatternFill(start_color='F09456',
                         end_color='F09456',
                         fill_type='solid')

orange_25 = PatternFill(start_color='F3AD7D',
                        end_color='F3AD7D',
                        fill_type='solid')

orange_275 = PatternFill(start_color='F8CBAD',
                         end_color='F8CBAD',
                         fill_type='solid')

orange_275 = PatternFill(start_color='F8CBAD',
                         end_color='F8CBAD',
                   fill_type='solid')

yellow_3 = PatternFill(start_color='FFC000',
                       end_color='FFC000',
                       fill_type='solid')

yellow_325 = PatternFill(start_color='FFDD71',
                         end_color='FFDD71',
                         fill_type='solid')

yellow_35 = PatternFill(start_color='FFE699',
                        end_color='FFE699',
                        fill_type='solid')

yellow_375 = PatternFill(start_color='FFF2CC',
                         end_color='FFF2CC',
                         fill_type='solid')

green_4 = PatternFill(start_color='92D050',
                      end_color='92D050',
                      fill_type='solid')


def color_cells(row_start, row_count, column):
    for i in range(row_start, row_count):
        cell_value = sheet.cell(row=i, column=column).value
        print(i, cell_value)
           try:
                if cell_value == 1:
                    sheet.cell(row=i, column=column).fill = red_1
                elif 1 < cell_value <= 1.25:
                    sheet.cell(row=i, column=column).fill = red_125
                elif 1.25 < cell_value <= 1.5:
                    sheet.cell(row=i, column=column).fill = red_15
                elif 1.5 < cell_value <= 1.75:
                    sheet.cell(row=i, column=column).fill = red_175
                elif 1.75 < cell_value <= 2:
                    sheet.cell(row=i, column=column).fill = orange_2
                elif 2 < cell_value <= 2.25:
                    sheet.cell(row=i, column=column).fill = orange_225
                elif 2.25 < cell_value <= 2.5:
                    sheet.cell(row=i, column=column).fill = orange_25
                elif 2.5 < cell_value <= 2.75:
                    sheet.cell(row=i, column=column).fill = orange_275
                elif 2.75 < cell_value <= 3:
                    sheet.cell(row=i, column=column).fill = yellow_3
                elif 3 < cell_value <= 3.25:
                    sheet.cell(row=i, column=column).fill = yellow_325
                elif 3.25 < cell_value <= 3.5:
                    sheet.cell(row=i, column=column).fill = yellow_35
                elif 3.5 < cell_value <= 3.75:
                    sheet.cell(row=i, column=column).fill = yellow_375
                elif 3.75 < cell_value <= 4:
                    sheet.cell(row=i, column=column).fill = green_4
            except BaseException:
                continue


row_header = 1
wb_sheets = wb.sheetnames
for sheet in wb_sheets:
    sheet = wb[sheet]
    numeric_columns = []
    print(sheet)
    for cell in sheet[1]:
        try:
            if cell_value.isnumeric():
                numeric_columns.append(cell.column)
        except BaseException:
            continue
    for col in numeric_columns:
        row_count = 0
        row_count = sheet.max_row + 1
        try:
            color_cells(row_header + 1, row_count, col)
        except BaseException:
            continue
    print(columns)
wb.save('C:\\Document-COLORED.xlsx')
