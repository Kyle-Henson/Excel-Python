import openpyxl
from openpyxl.styles import PatternFill
from _datetime import datetime, timedelta

# Colors time inside work hours green and outside yellow.
# Also converts for daylight savings time

wb = openpyxl.load_workbook('C:\\Document.xlsx'
                            , data_only=True)

green_fill = PatternFill(start_color='5db26b',
                   end_color='5db26b',
                   fill_type='solid')

yellow_fill = PatternFill(start_color='ffff00',
                   end_color='ffff00',
                   fill_type='solid')


def color_cells(row_start, row_count, column):
    for i in range(row_start, row_count):
        time = sheet.cell(row=i, column=column).value
        print(i, time)
        try:
            if 8 <= time.hour <= 18:
                sheet.cell(row=i, column=column).fill = green_fill
            else:
                sheet.cell(row=i, column=column).fill = yellow_fill
        except:
            continue


# Begins March 10, 2AM Ends November 3rd, 2AM. -4 instead of -5.
edt_start_str = "2019-03-10 02:00:00"
edt_start_obj = datetime.strptime(edt_start_str, '%Y-%m-%d %H:%M:%S')
edt_end_str = "2019-11-3 02:00:00"
edt_end_obj = datetime.strptime(edt_end_str, '%Y-%m-%d %H:%M:%S')

def convert_edt(row_start, row_count, column):
    print(column)
    for i in range(row_start, row_count):
        try:
            time = sheet.cell(row=i, column=column).value
            if edt_start_obj <= time <= edt_end_obj:
                edt_time = time - timedelta(hours=1)
                sheet.cell(row=i, column=column).value = edt_time
            else:
                print("")
        except:
            continue


row_header = 6
wb_sheets = wb.sheetnames
for sheet in wb_sheets:
    sheet = wb[sheet]
    EST_columns = []
    print(sheet)
    for cell in sheet[6]:
        try:
            if ('EST') in cell.value:
                EST_columns.append(cell.column)
        except:
            continue
    for col in EST_columns:
        row_count = 0
        row_count = sheet.max_row + 1
        try:
            convert_edt(row_header+1, row_count, col)
            color_cells(row_header + 1, row_count, col)
        except:
            continue
    print(EST_columns)
wb.save('C:\\Document-Colored.xlsx')
